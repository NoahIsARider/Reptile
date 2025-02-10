import pymysql
from openpyxl import load_workbook

# 连接到 MySQL 服务器
mydb = pymysql.connect(
    host="localhost",  # 你的 MySQL 主机地址
    user="root",  # 你的 MySQL 用户名
    password="050630zfyn"  # 你的 MySQL 密码
)

# 创建游标对象
mycursor = mydb.cursor()

# 创建数据库
mycursor.execute("CREATE DATABASE IF NOT EXISTS A24_movies")
mycursor.execute("USE A24_movies")

# 创建数据表
mycursor.execute("""
CREATE TABLE IF NOT EXISTS movies (
    id INT AUTO_INCREMENT PRIMARY KEY,
    chinese_name VARCHAR(255),
    english_name VARCHAR(255),
    rating VARCHAR(20),
    director VARCHAR(255),
    actors VARCHAR(255),
    movie_type VARCHAR(255),
    country VARCHAR(255),
    year VARCHAR(10)
)
""")

# 加载 Excel 文件
workbook = load_workbook(r"C:\Users\Lenovo\Desktop\movies.xlsx")
sheet = workbook.active

# 获取表头索引
header = [cell.value for cell in sheet[1]]
chinese_name_index = None
english_name_index = None
rating_index = None
director_index = None
actors_index = None
movie_type_index = None
country_index = None
year_index = None

for index, value in enumerate(header):
    if "电影名字" == value:
        chinese_name_index = index
    elif "评分" == value:
        rating_index = index
    elif "导演" == value:
        director_index = index
    elif "主演" == value:
        actors_index = index
    elif "类型" == value:
        movie_type_index = index
    elif "制片国家/地区" == value:
        country_index = index
    elif "年份" == value:
        year_index = index

# 假设电影名字列中的英文名在中文名后面用空格分隔
for row in sheet.iter_rows(min_row=2, values_only=True):
    movie_name = row[chinese_name_index] if chinese_name_index is not None else None
    if movie_name:
        space_index = movie_name.find(' ')
        if space_index != -1:
            chinese_name = movie_name[:space_index]
            english_name = movie_name[space_index + 1:]
        else:
            chinese_name = movie_name
            english_name = ""
    else:
        chinese_name = ""
        english_name = ""

    rating = row[rating_index] if rating_index is not None else None
    director = row[director_index] if director_index is not None else None
    actors = row[actors_index] if actors_index is not None else None
    movie_type = row[movie_type_index] if movie_type_index is not None else None
    country = row[country_index] if country_index is not None else None
    year = row[year_index] if year_index is not None else None

    sql = "INSERT INTO movies (chinese_name, english_name, rating, director, actors, movie_type, country, year) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
    val = (chinese_name, english_name, rating, director, actors, movie_type, country, year)
    mycursor.execute(sql, val)

# 提交事务
mydb.commit()

# 关闭游标和连接
mycursor.close()
mydb.close()